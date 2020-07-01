# import json
from openpyxl import Workbook
import datetime
from templates import days, Registrars, Consultants, weekdays

wb = Workbook()



titles= ["Date","Week","AM/PM","Room", "Clinic","Weeks","Clinician","GRADE"]
holsCons = []
holsReg = []
cancelList = []
consCall = []

def setupWeek(currentDate, dateFormatted):
    ws = wb.active
    ws.title = dateFormatted
    next_row = 2

    for day in days:

        # currentCol = 1
        # next_row = 2
        #set up lunar day and week.
        lunarWeek = int((currentDate.day -1)/7)+1
        weekday = currentDate.strftime("%w")
        # ws.active = int(weekday)+1
        # each day has it's own section 9 cols apart

        currentCol = (int(weekday) -1) *9

        #Who is on call?
        conOnCall = ""
        for con,dat in consCall:
            if weekday in dat:
                print("{} is on call day {}".format(con, weekday))
                conOnCall = con
                break


        #titles
        for i,title in enumerate(titles):
            ws.cell(row=1, column=currentCol +i+1).value = title

        ws.cell(row = next_row, column =1 + currentCol).value = day["DAY"]+day["AMPM"]

        next_row += 1


        if weekday in [0,6]:
            currentDate += datetime.timedelta(day=1)
            # print("weekend")
            continue


        #create the day - AM first then PM
        for clinic in day:
            if clinic in ["AMPM","Weeks","DAY"]:
                continue

            if lunarWeek not in day[clinic][2]:
                # print("Not on this week")
                continue
            name = day[clinic][0]
            grade = day[clinic][1]
            onCallStatus = day[clinic][3]
            # print(onCallStatus)
            if grade == "CONS":
                if conOnCall == name and onCallStatus != "KEEP":
                    if onCallStatus =="DROP":
                        cancelList.append("{} is on call on day {}  so {} is not on".format(name, weekday, clinic))
                        continue

                if onCallStatus == "SPECIALIN" and conOnCall != "KB":# when KB is not on call don't put in move of this clinic
                    print(onCallStatus)
                    cancelList.append("{} is on call on day {}  so {} is moved to the AM".format(name, weekday, clinic))
                    continue

                if onCallStatus == "SPECIALOUT" and conOnCall == "KB":# when KB is on call do remove this clinic
                    print(onCallStatus)
                    cancelList.append("{} is on call on day {}  so {} is moved from the PM".format(name, weekday, clinic))
                    continue

                if conOnCall != name and onCallStatus == "MOVE":# only put these clinics on when consultant on call
                    continue

                if isAvailable(name, weekday,holsCons) == False and "1088AA" not in clinic:
                    cancelList.append("{} is away on day {}  so {} is not on".format(name, weekday, clinic))
                    continue




            elif grade == "REG":
                for reg in holsReg:#is this teams reg on hols
                    if name in reg[0] and reg[0][-3:] == "REG":#tuple of name and dates- do the cons and reg match?
                        name = reg[0] # correct team so name of the registrar is now the name we're going to search for in holsReg
                        break
                if isAvailable(name, weekday,holsReg) == False:
                    cancelList.append("{} is unavailable on day {}  so {} is not on".format(name, weekday, clinic))
                    continue


            elif grade == "StG1" or grade == "StG2":
                for reg in holsReg:# check who is unavailable
                    if reg[0][-4:] == grade:
                        name = reg[0]
                        break
                if isAvailable(name, weekday,holsReg) == False:
                    cancelList.append("{} is unavailable on day {}  so {} is not on".format(name, weekday, clinic))
                    continue








            ws.cell(row=next_row, column = 5+ currentCol).value = clinic

            ws.cell(row=next_row, column = 1+ currentCol).value = currentDate.strftime('%d/%m/%Y')
            ws.cell(row=next_row, column = 2+ currentCol).value = lunarWeek


            ws.cell(row=next_row, column = 7+ currentCol).value = day[clinic][0]
            ws.cell(row=next_row, column = 8+ currentCol).value = day[clinic][1]


            next_row +=1



        next_row+=1
        #if finished PM, reset to next day (could use other sheet if wanted)
        if day["AMPM"] == "PM":
            currentDate += datetime.timedelta(days=1)
            # wb.create_sheet("{}".format(day["DAY"]))
            next_row = 2




    filename = "{}.xlsx".format(dateFormatted)
    # with open("{}_cancel.txt".format(dateFormatted),'w') as f:
    #     for i,can in enumerate(cancelList,25):
    #             f.write(can+'\n')
    #             ws.cell(row=i, column = 1).value = can
    #             print(can)

    wb.save(filename=filename)

def isAvailable(name, weekday, ClinicianList):
    for d in ClinicianList:
        con,dat = d
        if con == name:
            if str(weekday) in dat:
                return False
    return True




def getDate():
    day = int (input("Which date - use number 1 -31 "))
    month = int(input("Which month - use number 1-12 "))

    try:
        startDate = datetime.datetime(2020,month, day)
        if startDate.strftime("%A")!= "Monday":
            print("That is not a Monday")
            return None
        return startDate
    except:
        print("Sorry there is something wrong with those dates")
        return None

def getHotReg():
    while True:
        for i,reg in enumerate(Registrars):
            if i == 0:
                continue
            print(i,reg)
        try:
            hot = int(input("Who is hot reg?"))
        except ValueError:
            print("Please enter a valid number")
            continue
        if hot >=1 and hot <= len(Registrars):
            return Registrars[hot]
        print("Not one of the options, please try again.")


def getWeekendOnCall():
    while True:
        while True:
            for i,reg in enumerate(Registrars):
                print(i,reg)
            try:
                print("Type 0 if weekend reg does not need Monday off -eg locum")
                weekend = int(input("Who is weekend reg on call before the Monday?"))
                if weekend >=0 and weekend <= len(Registrars):
                    return weekend
                print("Not one of the options, please try again.")
                continue
            except ValueError:
                print("Please enter a valid number")
                continue


def leaveDays():
    while True:
        for d in weekdays:
            print(d)
        whichDay = None
        while not whichDay:
            print("Using the numbers key is the days of the week")
            print("that the consultant is unavailable, then press Enter.")
            print("E.g. if just Monday type 1, if all week type 12345.")
            try:
                whichDay = (input(""))
                if len(whichDay) <1 or len(whichDay)>5:
                    raise ValueError
                for f in whichDay:
                    if f not in "12345":
                        raise ValueError
                return (whichDay)
            except ValueError:
                print("Please try again - only numbers 12345")
                continue

def YorN(question):
    while True:
        print(question)
        answer = input("Input y or n only please:").lower()
        if answer != "y" and answer !="n":
            continue
        else:
            return answer

def getLeave(leaveList, leaveGroup,groupName):
    while True:
        for i,cons in enumerate(leaveGroup,1):
            print(i,cons)
        print("Type 0 for no more {} on leave".format(groupName))
        print("")
        try:
            leave = int(input("Which {} is on leave? ".format(groupName)))
            if leave == 0:
                break
            if leave >=1 and leave <= len(leaveGroup):
                print("")
                daysOff = leaveDays()
                leaveList.append((leaveGroup[leave-1], daysOff))
                spacer()
                for l in leaveList:
                    print("{} is off on days {}".format(l[0],l[1]) )
                spacer()
                again = YorN("Any more {} leave?".format(groupName))
                if again == "y":
                    continue
                else:
                    break
            else:
                raise ValueError

        except ValueError:
            print("Please enter a valid number")
            continue

def getOnCall():
    while True:
        for i,cons in enumerate(Consultants,1):
            print(i,cons)
        print("Type 0 for no more Consultants on call")
        print("")
        try:
            call = int(input("Which consultant is on call? "))
            if call == 0:
                break

            elif call >= 1 and call <= len(Consultants):
                chosen = Consultants[call-1]
                print("")
                if chosen in ["SS","RS","RGW"]:
                    print ("{} not doing consultant on call at JCUH\n".format(chosen))
                    raise ValueError

                callDays = leaveDays()
                consCall.append((chosen, callDays))
                spacer()
                again = YorN("Any more Consultants on call?")
                if again == "y":
                    continue
                else:
                    break

            else:
                raise ValueError

        except ValueError:
            print("Please enter a valid number")
            continue

    checkOnCall = ""
    for _,d in consCall:
        checkOnCall += d
    checkList = "".join(sorted(checkOnCall))
    if checkList == "12345":
        return True
    print("")
    print("Not all days covered correctly. Try again")
    consCall.clear()
    return False


def spacer():
    print("")
    print("*" * 10)
    print("")

def main():
    print("Ready to set up a timetable?")
    print("If you are looking for after 2020, contact Mr Lester")

    startDate = None
    while startDate == None:
        startDate = getDate()
    currentDate = startDate
    readable = currentDate.strftime('%d_%b_%Y')
    print("\n" + readable +"\n")

    if YorN("Is this the correct date?") == "n":
        return

    #start collection leave data
    spacer()
    hot = getHotReg()
    holsReg.append((hot,"12345"))#they are unavailable
    spacer()
    weekend = getWeekendOnCall()
    if weekend !=0:
        holsReg.append((Registrars[weekend],"1"))
    spacer()
    getLeave(holsCons, Consultants, "consultants")
    spacer()
    getLeave(holsReg, Registrars[1:], "registrars")
    spacer()

    while getOnCall() == False:
        pass



    print("")
    print("For the week {} ...".format(readable))
    print("")
    print("hot reg {}, weekend reg - Monday off {}".format(hot, Registrars[weekend]))
    print("")
    print("Consultants on call")
    print("")
    for call in consCall:
        print(call[0], call[1])
    print("")
    print("Consultant leave")
    print("")
    for cons in holsCons:
        print (cons[0], cons[1])
    print("")
    print("Registrar leave")
    print("")
    for reg in holsReg:
        print (reg[0], reg[1])
    spacer()

    if YorN("Is this correct?") == "y":
        print("Great- producing timetable")
        setupWeek(currentDate, readable)


    else:
        main()

if __name__ =="__main__":
    main()

